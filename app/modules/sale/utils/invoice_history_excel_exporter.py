import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class InvoiceHistoryExcelExporter:
    """
    Tiện ích chuyên biệt kết xuất dữ liệu chi tiết Hóa đơn ra tệp Excel
    """

    @staticmethod
    def export_detail(file_path: str, metadata: dict, items: list) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Chi tiết Hóa đơn"

            # === ĐỊNH NGHĨA PALETTE MÀU SẮC ENTERPRISE ===
            title_font = Font(name="Segoe UI", size=14, bold=True, color="1E293B")
            meta_font = Font(name="Segoe UI", size=10, bold=False, color="475569")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            footer_font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")

            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Màu xanh dương

            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            thin_border = Border(
                left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"),
                top=Side(style='thin', color="CBD5E1"), bottom=Side(style='thin', color="CBD5E1")
            )

            # 1. Dòng 1: Tiêu đề báo cáo gộp ô
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = "CHI TIẾT ẤN PHẨM HÓA ĐƠN BÁN HÀNG"
            title_cell.font = title_font
            title_cell.alignment = center_align
            ws.row_dimensions[1].height = 30

            # 2. Dòng 3 & Dòng 4: Thiết lập Metadata phân lớp rõ ràng qua tọa độ cứng
            ws['A3'] = f"Mã hóa đơn: {metadata['code']}"
            ws['A3'].font = meta_font
            ws['A3'].alignment = left_align

            created_time = metadata['created_at'].strftime('%d/%m/%Y %H:%M') if isinstance(metadata['created_at'],
                                                                                           datetime.datetime) else str(
                metadata['created_at'])
            ws['D3'] = f"Ngày bán: {created_time}"
            ws['D3'].font = meta_font
            ws['D3'].alignment = left_align

            status_text = "Hoàn thành" if metadata[
                                              'status'] == 'COMPLETED' else f"Đã hủy (Lý do: {metadata.get('cancel_reason', '')})"
            ws['A4'] = f"Trạng thái: {status_text}"
            ws['A4'].font = Font(name="Segoe UI", size=10, bold=True,
                                 color="EF4444" if metadata['status'] == 'CANCELLED' else "059669")
            ws['A4'].alignment = left_align

            method_text = "Tiền mặt" if metadata['payment_method'] == 'CASH' else "Chuyển khoản"
            ws['D4'] = f"Hình thức: {method_text}"
            ws['D4'].font = meta_font
            ws['D4'].alignment = left_align

            # 3. Dòng 6: Thiết lập Header danh sách hàng hóa
            headers = ["STT", "Tên Sản Phẩm", "ĐVT", "Số Lượng", "Đơn Giá (VNĐ)", "Thành Tiền (VNĐ)"]
            header_row = 6
            ws.row_dimensions[header_row].height = 25

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # 4. Từ Dòng 7 trở đi: Vòng lặp đổ dữ liệu chi tiết
            current_row = header_row + 1
            for index, item in enumerate(items, 1):
                ws.row_dimensions[current_row].height = 20

                ws.cell(row=current_row, column=1, value=index).alignment = center_align
                ws.cell(row=current_row, column=2, value=item['product_name']).alignment = left_align
                ws.cell(row=current_row, column=3, value=item['unit_name']).alignment = center_align

                qty_cell = ws.cell(row=current_row, column=4, value=int(item['quantity']))
                qty_cell.alignment = center_align
                qty_cell.number_format = '#,##0'

                price_cell = ws.cell(row=current_row, column=5, value=float(item['unit_price']))
                price_cell.alignment = right_align
                price_cell.number_format = '#,##0'

                total_cell = ws.cell(row=current_row, column=6, value=float(item['total_price']))
                total_cell.alignment = right_align
                total_cell.number_format = '#,##0'

                # Áp viền mỏng cho toàn bộ các ô dữ liệu
                for col_num in range(1, 7):
                    ws.cell(row=current_row, column=col_num).border = thin_border

                current_row += 1

            # 5. Dòng kế tiếp: Ghi nhận Tổng cộng hóa đơn tài chính
            ws.row_dimensions[current_row].height = 24
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)

            footer_label = ws.cell(row=current_row, column=1)
            footer_label.value = "TỔNG CỘNG KHÁCH TRẢ"
            footer_label.font = footer_font
            footer_label.alignment = right_align

            footer_val = ws.cell(row=current_row, column=6)
            footer_val.value = float(metadata['final_amount'])
            footer_val.font = Font(name="Segoe UI", size=11, bold=True, color="EF4444")
            footer_val.number_format = '#,##0'
            footer_val.alignment = right_align

            # Đóng khung viền đầy đủ cho dòng tổng cộng
            for col_num in range(1, 7):
                ws.cell(row=current_row, column=col_num).border = thin_border

            # 6. Định dạng độ rộng tối ưu cho các cột chống tràn chữ (###)
            col_widths = {'A': 6, 'B': 42, 'C': 12, 'D': 12, 'E': 18, 'F': 22}
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            # 7. Lưu file vật lý xuống bộ nhớ máy tính
            wb.save(file_path)
            return True

        except Exception as e:
            raise Exception(f"Có lỗi khi xuất Excel chi tiết hóa đơn: {str(e)}")