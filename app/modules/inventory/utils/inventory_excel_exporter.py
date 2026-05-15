import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class InventoryExcelExporter:
    """
    Tiện ích chuyên biệt để xuất Báo cáo Tồn kho ra file Excel.
    """

    @staticmethod
    def export(data: list, file_path: str) -> bool:
        try:
            # Khởi tạo Workbook và Worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Báo Cáo Tồn Kho"

            # === ĐỊNH NGHĨA CÁC STYLE CHO EXCEL ===
            title_font = Font(size=16, bold=True, color="1E293B")
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            warning_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")

            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"),
                top=Side(style='thin', color="CBD5E1"), bottom=Side(style='thin', color="CBD5E1")
            )

            # Viết Tiêu đề báo cáo
            ws.merge_cells('A1:H1')
            cell_title = ws['A1']
            cell_title.value = f"BÁO CÁO TỒN KHO - Cập nhật ngày {datetime.datetime.now().strftime('%d/%m/%Y')}"
            cell_title.font = title_font
            cell_title.alignment = center_align

            # Viết Header cho Bảng
            headers = ["STT", "Mã SKU", "Tên Sản Phẩm", "ĐVT", "Tồn Kho", "Định Mức", "Giá Vốn (VNĐ)", "Tổng Giá Trị (VNĐ)"]
            ws.append([])
            ws.append(headers)

            header_row = 3
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Đổ Dữ liệu và Áp dụng Style
            total_qty = 0
            total_val = 0
            current_row = header_row + 1

            for index, row_data in enumerate(data, 1):
                qty = int(row_data['quantity'])
                min_stock = int(row_data['min_stock'])
                cost_price = float(row_data['cost_price']) if row_data['cost_price'] else 0
                val = float(row_data['total_value']) if row_data['total_value'] else 0

                total_qty += qty
                total_val += val

                ws.append([
                    index,
                    row_data['sku'],
                    row_data['product_name'],
                    row_data['unit_name'],
                    qty,
                    min_stock,
                    cost_price,
                    val
                ])

                for col_num in range(1, 9):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.border = thin_border
                    if col_num in [1, 2, 4, 5, 6]:
                        cell.alignment = center_align
                    if col_num in [5, 6, 7, 8]:
                        cell.number_format = '#,##0'

                # Bôi vàng cảnh báo
                if qty <= min_stock:
                    for col_num in range(1, 9):
                        ws.cell(row=current_row, column=col_num).fill = warning_fill

                current_row += 1

            # Dòng Tổng Cộng ở cuối bảng
            ws.merge_cells(f'A{current_row}:D{current_row}')
            footer_cell = ws.cell(row=current_row, column=1)
            footer_cell.value = "TỔNG CỘNG HÀNG HÓA"
            footer_cell.font = Font(bold=True, color="1E293B")
            footer_cell.alignment = Alignment(horizontal="right", vertical="center")

            ws.cell(row=current_row, column=5).value = total_qty
            ws.cell(row=current_row, column=8).value = total_val

            for col_num in range(1, 9):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = thin_border
                cell.font = Font(bold=True, color="EF4444")
                if col_num in [5, 8]:
                    cell.number_format = '#,##0'

            # Căn chỉnh độ rộng cột
            col_widths = {'A': 5, 'B': 15, 'C': 40, 'D': 10, 'E': 12, 'F': 12, 'G': 18, 'H': 22}
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            # Lưu file
            wb.save(file_path)
            return True

        except Exception as e:
            raise Exception(f"Có lỗi khi tạo file Excel: {str(e)}")