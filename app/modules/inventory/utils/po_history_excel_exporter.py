import traceback
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.modules.inventory.dtos.po_history_dto import PurchaseOrderMasterDTO, PurchaseOrderDetailDTO


class PoHistoryExcelExporter:

    @staticmethod
    def export_detail(file_path: str, po_master: PurchaseOrderMasterDTO,
                      po_items: List[PurchaseOrderDetailDTO]) -> bool:
        """
        Xuất thông tin chi tiết của MỘT phiếu nhập cụ thể ra file Excel.
        - file_path: Đường dẫn lưu file (.xlsx)
        - po_master: DTO chứa thông tin chung (Header) của phiếu nhập
        - po_items: Danh sách DTO chứa các mặt hàng trong phiếu
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"ChiTiet_{po_master.code}"

            # Đảm bảo hiển thị đường lưới ô (Gridlines)
            ws.views.sheetView[0].showGridLines = True

            # ---- ĐỊNH NGHĨA STYLE (Màu sắc nhã nhặn) ----
            font_title = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
            font_section = Font(name="Segoe UI", size=11, bold=True, color="334155")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_body = Font(name="Segoe UI", size=11)
            font_total = Font(name="Segoe UI", size=11, bold=True, color="B91C1C")

            fill_header = PatternFill(start_color="334155", end_color="334155",
                                      fill_type="solid")  # Màu xám xanh slate đậm
            fill_cancelled = PatternFill(start_color="FEE2E2", end_color="FEE2E2",
                                         fill_type="solid")  # Màu đỏ nhạt cho phiếu hủy

            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")

            # Định nghĩa viền ô mảnh màu xám
            thin_border_side = Side(border_style="thin", color="CBD5E1")
            border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side,
                                 bottom=thin_border_side)

            # ---- VẼ PHẦN TIÊU ĐỀ & THÔNG TIN CHUNG (META DATA) ----
            ws["A1"] = "BÁO CÁO CHI TIẾT PHIẾU NHẬP KHO"
            ws["A1"].font = font_title

            # Thông tin chung sắp xếp theo các dòng
            meta_data = [
                ("Mã phiếu nhập:", po_master.code),
                ("Ngày nhập kho:", po_master.created_at.strftime("%d/%m/%Y %H:%M")),
                ("Nhà cung cấp:", po_master.supplier_name),
                ("Trạng thái phiếu:", "ĐÃ HỦY" if po_master.status == "CANCELLED" else "HOÀN THÀNH"),
                ("Ghi chú nhập:", po_master.note or "---")
            ]

            # Nếu phiếu đã hủy, bổ sung thêm dòng lý do hủy
            if po_master.status == "CANCELLED":
                meta_data.append(("Lý do hủy phiếu:", po_master.cancel_reason or "---"))

            current_row = 3
            for label, val in meta_data:
                ws.cell(row=current_row, column=1, value=label).font = font_section
                cell_val = ws.cell(row=current_row, column=2, value=val)
                cell_val.font = font_body

                # Highlight đặc biệt nếu trạng thái là ĐÃ HỦY
                if label == "Trạng thái phiếu:" and po_master.status == "CANCELLED":
                    cell_val.font = Font(name="Segoe UI", size=11, bold=True, color="991B1B")
                    cell_val.fill = fill_cancelled
                current_row += 1

            # ---- VẼ BẢNG CHI TIẾT MẶT HÀNG (TABLE ITEMS) ----
            current_row += 1  # Cách ra 1 dòng trống trống trải
            headers = ["STT", "Mã SKU", "Tên Sản Phẩm / Ấn Phẩm", "Đơn Vị Tính", "Số Lượng", "Đơn Giá Nhập",
                       "Thành Tiền"]

            # Đổ dữ liệu Header dòng bảng
            for col_idx, header_text in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header_text)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center if col_idx in [1, 2, 4] else (
                    align_right if col_idx in [5, 6, 7] else align_left)
                ws.row_dimensions[current_row].height = 25

            # Đổ dữ liệu Items thân bảng
            start_table_row = current_row + 1
            for stt, item in enumerate(po_items, start=1):
                current_row += 1
                row_data = [
                    stt,
                    item.sku,
                    item.product_name,
                    item.unit_name,
                    item.quantity,
                    item.unit_price,
                    item.total_price
                ]

                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = font_body
                    cell.border = border_cell
                    ws.row_dimensions[current_row].height = 20

                    # Cấu hình định dạng căn lề và format tiền tệ/số lượng theo từng cột
                    if col_idx in [1, 4]:  # STT, ĐVT
                        cell.alignment = align_center
                    elif col_idx == 2:  # SKU
                        cell.alignment = align_center
                    elif col_idx == 3:  # Tên SP
                        cell.alignment = align_left
                    elif col_idx == 5:  # Số lượng
                        cell.alignment = align_right
                        cell.number_format = '#,##0'
                    elif col_idx in [6, 7]:  # Đơn giá, Thành tiền
                        cell.alignment = align_right
                        cell.number_format = '#,##0'

            # ---- DÒNG TỔNG CỘNG (TOTAL BILL FOOTER) ----
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            total_label_cell = ws.cell(row=current_row, column=1, value="TỔNG GIÁ TRỊ PHIẾU NHẬP:")
            total_label_cell.font = font_section
            total_label_cell.alignment = align_right

            # Sử dụng công thức SUM của Excel thay vì ghi giá trị cứng (Chuẩn nghiệp vụ)
            total_value_cell = ws.cell(row=current_row, column=7, value=f"=SUM(G{start_table_row}:G{current_row - 1})")
            total_value_cell.font = font_total
            total_value_cell.alignment = align_right
            total_value_cell.number_format = '#,##0 "VND"'
            ws.row_dimensions[current_row].height = 25

            # ---- TỰ ĐỘNG CÂN ĐỐI ĐỘ RỘNG CÁC CỘT (AUTO-FIT WIDTH) ----
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row < start_table_row - 1:  # Bỏ qua các dòng Meta phía trên để tránh loãng độ rộng
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                # Cộng thêm 4 đơn vị padding an toàn chống lỗi hiển thị dấu "###"
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # Đặc quyền chỉnh tay độ rộng cột Tên sản phẩm cho rộng rãi
            ws.column_dimensions['C'].width = 35

            # Lưu Workbook ra file thực tế
            wb.save(file_path)
            return True

        except Exception as e:
            print(f"Lỗi xuất file Excel chi tiết phiếu nhập: {str(e)}")
            traceback.print_exc()
            return False