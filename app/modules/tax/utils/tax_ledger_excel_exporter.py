
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class TaxLedgerExcelExporter:
    """
    Tiện ích chuyên biệt kết xuất dữ liệu Sổ cái quyết toán thuế ra tệp Excel chuẩn chỉnh,
    bảo toàn tính toàn vẹn của dữ liệu số học kế toán.
    """

    @staticmethod
    def export_ledger(file_path: str, metadata: dict, items: list) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Sổ cái Thuế {metadata['year']}"

            # === ĐỊNH NGHĨA PALETTE MÀU SẮC ENTERPRISE ===
            title_font = Font(name="Segoe UI", size=14, bold=True, color="0F172A")
            meta_font = Font(name="Segoe UI", size=10, bold=False, color="334155")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            footer_font = Font(name="Segoe UI", size=11, bold=True, color="0F172A")

            # Sử dụng màu xanh Navy đậm sang trọng cho tiêu đề quyết toán
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            thin_border = Border(
                left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"),
                top=Side(style='thin', color="CBD5E1"), bottom=Side(style='thin', color="CBD5E1")
            )

            # 1. Dòng 1: Tiêu đề báo cáo gộp ô
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = f"SỔ CÁI QUYẾT TOÁN THUẾ NĂM {metadata['year']}"
            title_cell.font = title_font
            title_cell.alignment = center_align
            ws.row_dimensions[1].height = 30

            # 2. Dòng 3, 4, 5: Thiết lập Metadata cấu hình đóng băng lịch sử
            ws['A3'] = f"Mức miễn thuế cơ sở: {metadata['threshold']:,.0f} VND"
            ws['A3'].font = meta_font
            ws['A3'].alignment = left_align

            ws['D3'] = f"Phương pháp tính thuế TNCN: {metadata['method_display']}"
            ws['D3'].font = meta_font
            ws['D3'].alignment = left_align

            ws['A4'] = f"Thuế suất GTGT: {metadata['vat_percent']}%"
            ws['A4'].font = meta_font
            ws['A4'].alignment = left_align

            ws['D4'] = f"Thuế suất TNCN: {metadata['pit_percent']}%"
            ws['D4'].font = meta_font
            ws['D4'].alignment = left_align

            is_closed = "CLOSED" in metadata['status'] or "🔒" in metadata['status']
            status_clean = "🔒 Đã khóa sổ" if is_closed else "⚠️ Bản nháp (Mở duyệt)"
            ws['A5'] = f"Trạng thái sổ cái: {status_clean}"
            ws['A5'].font = Font(name="Segoe UI", size=10, bold=True, color="059669" if is_closed else "D97706")
            ws['A5'].alignment = left_align

            ws['D5'] = f"Ngày quyết toán khóa sổ: {metadata['finalized_at']}"
            ws['D5'].font = meta_font
            ws['D5'].alignment = left_align

            # 3. Dòng 7: Tiêu đề các cột dữ liệu số
            headers = ["Kỳ kế toán", "Doanh thu gộp", "Tổng chi phí", "Thuế GTGT", "Thuế TNCN", "Tổng thuế phải nộp"]
            header_row = 7
            ws.row_dimensions[header_row].height = 25

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # 4. Từ Dòng 8 trở đi: Đổ dữ liệu 12 tháng từ mảng DTO
            current_row = header_row + 1
            for item in items:
                ws.row_dimensions[current_row].height = 20

                ws.cell(row=current_row, column=1, value=f"Tháng {item.month}").alignment = center_align

                # Trích xuất và format số đại số kế toán
                vals = [item.revenue, item.cost, item.vat_amount, item.pit_amount, (item.vat_amount + item.pit_amount)]
                for col_idx, val in enumerate(vals, 2):
                    cell = ws.cell(row=current_row, column=col_idx, value=float(val))
                    cell.alignment = right_align
                    cell.number_format = '#,##0'

                for col_num in range(1, 7):
                    ws.cell(row=current_row, column=col_num).border = thin_border

                current_row += 1

            # 5. Dòng kết tiếp: Tính toán tổng cộng lũy kế toàn năm
            ws.row_dimensions[current_row].height = 24
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)

            footer_label = ws.cell(row=current_row, column=1)
            footer_label.value = "TỔNG CỘNG QUYẾT TOÁN"
            footer_label.font = footer_font
            footer_label.alignment = right_align

            # Thực hiện phép tính cộng dồn an toàn
            total_vat = sum(item.vat_amount for item in items)
            total_pit = sum(item.pit_amount for item in items)
            total_tax = total_vat + total_pit

            ws.cell(row=current_row, column=4, value=float(total_vat)).number_format = '#,##0'
            ws.cell(row=current_row, column=5, value=float(total_pit)).number_format = '#,##0'

            tot_tax_cell = ws.cell(row=current_row, column=6, value=float(total_tax))
            tot_tax_cell.font = Font(name="Segoe UI", size=11, bold=True, color="EF4444")
            tot_tax_cell.number_format = '#,##0'

            for col_num in range(4, 7):
                ws.cell(row=current_row, column=col_num).font = footer_font
                ws.cell(row=current_row, column=col_num).alignment = right_align

            for col_num in range(1, 7):
                ws.cell(row=current_row, column=col_num).border = thin_border

            # 6. Đặt độ rộng tối ưu để chống lỗi tràn ký tự (###)
            col_widths = {'A': 15, 'B': 22, 'C': 22, 'D': 18, 'E': 18, 'F': 24}
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            wb.save(file_path)
            return True

        except Exception as e:
            raise Exception(f"Có lỗi khi xuất Excel chi tiết sổ cái thuế: {str(e)}")